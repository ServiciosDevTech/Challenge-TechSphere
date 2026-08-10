from __future__ import annotations

import json
import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import get_settings

logger = logging.getLogger(__name__)


def _dataset_root() -> Path:
    settings = get_settings()
    # Prefer Challenge-TechSphere/data/dataset (junction to artifacts)
    repo_data = Path(__file__).resolve().parents[2] / "data" / "dataset"
    if (repo_data / "dataset_final.xlsx").exists():
        return repo_data
    # Fallback: sibling ParticipantArtifacts-main
    sibling = Path(__file__).resolve().parents[3] / "ParticipantArtifacts-main" / "dataset"
    if (sibling / "dataset_final.xlsx").exists():
        return sibling
    configured = Path(getattr(settings, "dataset_dir", "") or "")
    if configured and (configured / "dataset_final.xlsx").exists():
        return configured
    return repo_data


def _read_sheet(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        logger.warning("Dataset file missing: %s", path)
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["result"] if "result" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows))]
        out: list[dict[str, Any]] = []
        for row in rows:
            item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            out.append(item)
        return out
    finally:
        wb.close()


def _parse_jsonish(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (list, dict)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text


@lru_cache(maxsize=1)
def load_dataset_bundle() -> dict[str, Any]:
    root = _dataset_root()
    dialogs = _read_sheet(root / "dataset_final.xlsx")
    trajectories = _read_sheet(root / "trayectorias_postop_silver.xlsx")
    clinical = _read_sheet(root / "perfiles_clinicos_pacientes_silver_contest.xlsx")
    demo = _read_sheet(root / "perfiles_pacientes_co.xlsx")

    clinical_by_id = {str(r["paciente_id"]): r for r in clinical if r.get("paciente_id")}
    demo_by_id = {str(r["paciente_id"]): r for r in demo if r.get("paciente_id")}
    traj_by_id = {str(r["trayectoria_id"]): r for r in trajectories if r.get("trayectoria_id")}

    cases: dict[str, dict[str, Any]] = {}
    for row in dialogs:
        caso_id = str(row.get("caso_id") or "")
        if not caso_id:
            continue
        case = cases.setdefault(
            caso_id,
            {
                "caso_id": caso_id,
                "paciente_id": str(row.get("paciente_id") or ""),
                "dia_postop": row.get("dia_postop"),
                "label_ground_truth": str(row.get("label_ground_truth") or "").lower(),
                "turns": {"capa1_limpia": [], "capa2_ruidosa": []},
            },
        )
        capa = str(row.get("capa") or "capa1_limpia")
        if capa not in case["turns"]:
            case["turns"][capa] = []
        case["turns"][capa].append(
            {
                "turno_idx": row.get("turno_idx"),
                "hablante": row.get("hablante"),
                "texto": row.get("texto") or "",
                "dialogo_id": row.get("dialogo_id"),
            }
        )

    for case in cases.values():
        for capa_turns in case["turns"].values():
            capa_turns.sort(key=lambda t: (t.get("turno_idx") is None, t.get("turno_idx") or 0))

    patients: list[dict[str, Any]] = []
    for pid, clin in clinical_by_id.items():
        dem = demo_by_id.get(pid, {})
        patient_cases = [
            {
                "caso_id": c["caso_id"],
                "dia_postop": c["dia_postop"],
                "label_ground_truth": c["label_ground_truth"],
                "trayectoria_id": c["caso_id"].removeprefix("caso_")
                if c["caso_id"].startswith("caso_")
                else None,
            }
            for c in cases.values()
            if c["paciente_id"] == pid
        ]
        patient_cases.sort(key=lambda x: (x["dia_postop"] is None, x["dia_postop"] or 0))
        patients.append(
            {
                "paciente_id": pid,
                "nombre_completo": dem.get("nombre_completo") or pid,
                "ciudad": dem.get("ciudad"),
                "departamento": dem.get("departamento"),
                "eps": dem.get("eps"),
                "procedimiento": clin.get("procedimiento"),
                "fecha_cirugia": str(clin.get("fecha_cirugia") or ""),
                "edad": clin.get("edad"),
                "genero": clin.get("genero"),
                "comorbilidades": _parse_jsonish(clin.get("comorbilidades")),
                "modulo_synthea": clin.get("modulo_synthea"),
                "cases": patient_cases,
            }
        )

    patients.sort(key=lambda p: str(p.get("nombre_completo") or ""))

    return {
        "root": str(root),
        "patients": patients,
        "cases": cases,
        "trajectories": traj_by_id,
        "clinical": clinical_by_id,
        "demographics": demo_by_id,
        "stats": {
            "patients": len(patients),
            "cases": len(cases),
            "dialog_rows": len(dialogs),
            "trajectories": len(traj_by_id),
            "available": bool(cases),
        },
    }


def list_patients(limit: int | None = None) -> list[dict[str, Any]]:
    bundle = load_dataset_bundle()
    patients = bundle["patients"]
    if limit is not None:
        return patients[:limit]
    return patients


def list_procedures() -> list[str]:
    """Procedimientos únicos en perfiles clínicos del dataset."""
    patients = load_dataset_bundle()["patients"]
    seen: set[str] = set()
    out: list[str] = []
    for p in patients:
        raw = p.get("procedimiento")
        if raw is None:
            continue
        name = str(raw).strip()
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())
        out.append(name)
    out.sort(key=lambda s: s.casefold())
    return out


def get_case(caso_id: str) -> dict[str, Any] | None:
    return load_dataset_bundle()["cases"].get(caso_id)


def get_trajectory_for_case(caso_id: str) -> dict[str, Any] | None:
    traj_id = caso_id.removeprefix("caso_") if caso_id.startswith("caso_") else caso_id
    return load_dataset_bundle()["trajectories"].get(traj_id)


def resolve_call_context(
    *,
    paciente_id: str | None = None,
    caso_id: str | None = None,
    patient_name: str | None = None,
    procedure: str | None = None,
    dia_postop: int | None = None,
) -> dict[str, Any]:
    """
    Contexto para la llamada.
    - agent_context: lo que puede saber el agente al inicio (sin spoilear síntomas).
    - demo_script: guion oculto para el operador (trayectoria + label), NO enviar al LLM.
    """
    bundle = load_dataset_bundle()
    case = get_case(caso_id) if caso_id else None
    pid = str(paciente_id or (case or {}).get("paciente_id") or "")
    patient = next((p for p in bundle["patients"] if p["paciente_id"] == pid), None)

    if case is None and pid and dia_postop is not None:
        for c in bundle["cases"].values():
            if c["paciente_id"] == pid and int(c.get("dia_postop") or -1) == int(dia_postop):
                case = c
                caso_id = c["caso_id"]
                break

    if case is None and pid:
        # primer caso del paciente
        for p in bundle["patients"]:
            if p["paciente_id"] == pid and p["cases"]:
                caso_id = p["cases"][0]["caso_id"]
                case = get_case(caso_id)
                break

    traj = get_trajectory_for_case(caso_id) if caso_id else None
    name = (
        (patient or {}).get("nombre_completo")
        or patient_name
        or "Paciente Demo"
    )
    proc = (patient or {}).get("procedimiento") or procedure or "Apendicectomía"
    day = (case or {}).get("dia_postop") if case else dia_postop

    agent_context = {
        "patient_name": name,
        "patient_id": pid or None,
        "caso_id": caso_id,
        "procedure": proc,
        "dia_postop": day,
        "edad": (patient or {}).get("edad"),
        "genero": (patient or {}).get("genero"),
        "ciudad": (patient or {}).get("ciudad"),
        "eps": (patient or {}).get("eps"),
        # comorbilidades como contexto demográfico/clínico de fondo, no como síntomas actuales
        "comorbilidades": (patient or {}).get("comorbilidades"),
        "source": "challenge_dataset" if case or patient else "manual",
    }

    demo_script = None
    if traj or case:
        demo_script = {
            "label_ground_truth": (case or {}).get("label_ground_truth"),
            "arquetipo_trayectoria": (traj or {}).get("arquetipo_trayectoria"),
            "dolor_nrs": (traj or {}).get("dolor_nrs"),
            "fiebre_c": (traj or {}).get("fiebre_c"),
            "movilidad": (traj or {}).get("movilidad"),
            "herida": (traj or {}).get("herida"),
            "apetito": (traj or {}).get("apetito"),
            "sueno": (traj or {}).get("sueno"),
            "hint": _demo_hint(traj, (case or {}).get("label_ground_truth")),
        }

    return {
        "agent_context": agent_context,
        "demo_script": demo_script,
        "display_name": first_name(str(name)),
        "procedure": proc,
        "dia_postop": day,
        "paciente_id": pid or None,
        "caso_id": caso_id,
    }


def first_name(full_name: str) -> str:
    return full_name.strip().split()[0] if full_name.strip() else "paciente"


def _demo_hint(traj: dict[str, Any] | None, label: str | None) -> str:
    if not traj:
        return "Caso del dataset sin trayectoria asociada."
    parts = [
        f"Label esperado: {label or 'desconocido'}.",
        f"Dolor NRS={traj.get('dolor_nrs')}, fiebre={traj.get('fiebre_c')}°C.",
        f"Herida: {traj.get('herida')}; movilidad: {traj.get('movilidad')}.",
        "Interpreta este cuadro en lenguaje cotidiano; no leas el guion al agente.",
    ]
    return " ".join(str(p) for p in parts if p)


def patient_utterances_for_case(caso_id: str, capa: str = "capa1_limpia") -> list[str]:
    case = get_case(caso_id)
    if not case:
        return []
    turns = case["turns"].get(capa) or []
    return [
        str(t["texto"]).strip()
        for t in turns
        if str(t.get("hablante") or "").lower() in {"paciente", "patient"}
        and str(t.get("texto") or "").strip()
    ]


def iter_eval_cases(capa: str = "capa1_limpia") -> list[dict[str, Any]]:
    """Casos listos para eval offline (label + texto paciente)."""
    out = []
    for caso_id, case in load_dataset_bundle()["cases"].items():
        utterances = patient_utterances_for_case(caso_id, capa=capa)
        if not utterances:
            continue
        out.append(
            {
                "caso_id": caso_id,
                "paciente_id": case["paciente_id"],
                "dia_postop": case["dia_postop"],
                "label_ground_truth": case["label_ground_truth"],
                "patient_text": " ".join(utterances),
                "utterances": utterances,
            }
        )
    return out


def scenario_from_procedure(procedure: str | None) -> str | None:
    if not procedure:
        return None
    lower = procedure.lower()
    mapping = [
        (r"apendic", "Appendicitis"),
        (r"colecist|ves[ií]cula", "cholecystitis"),
        (r"mama|breast|cuello uterino|c[eé]rvix", "breast_cancer"),
        (r"colon|colorrect|rectal", "colorectal cancer"),
        (r"cadera|rodilla|artroplast|reemplazo|joint", "total joint replacement"),
    ]
    for pattern, folder in mapping:
        if re.search(pattern, lower):
            return folder
    return None
